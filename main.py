import json
import os
import re
import secrets
import sqlite3
import time
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import anthropic
import bcrypt
import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jose import jwt as jose_jwt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

load_dotenv()

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
MODEL_NAME = os.getenv("MODEL_NAME", "claude-sonnet-4-5")
PORT = int(os.getenv("PORT", "8000"))
DATABASE_PATH = os.getenv("DATABASE_PATH", "packinglist.db")
JWT_SECRET = os.getenv("JWT_SECRET", "")
LISTS_FILE = os.getenv("LISTS_FILE", "lists.json")  # migration source only
BASE_USER = os.getenv("BASE_USER", "michael")
BASE_PASS = os.getenv("BASE_PASS", "changeme")
INVITE_CODE = os.getenv("INVITE_CODE", "")  # required for registration when set

# Rate limit: max failed login attempts per IP
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300  # 5 minutes

# Persist a random JWT secret across restarts when not set via env
if not JWT_SECRET:
    _sp = Path(DATABASE_PATH).parent / ".jwt_secret"
    _sp.parent.mkdir(parents=True, exist_ok=True)
    if _sp.exists():
        JWT_SECRET = _sp.read_text().strip()
    else:
        JWT_SECRET = secrets.token_hex(32)
        _sp.write_text(JWT_SECRET)

app = FastAPI()

# ---------------------------------------------------------------------------
# Rate limiter (in-memory, resets on restart — fine for this scale)
# ---------------------------------------------------------------------------
_login_attempts: dict[str, list[float]] = {}  # ip -> [timestamps]


def _check_rate_limit(ip: str) -> None:
    now = time.time()
    attempts = _login_attempts.get(ip, [])
    # Prune old attempts outside the window
    attempts = [t for t in attempts if now - t < LOGIN_WINDOW_SECONDS]
    _login_attempts[ip] = attempts
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        raise HTTPException(
            status_code=429,
            detail=f"Too many login attempts. Try again in {LOGIN_WINDOW_SECONDS // 60} minutes.",
        )


def _record_failed_login(ip: str) -> None:
    _login_attempts.setdefault(ip, []).append(time.time())


def _clear_login_attempts(ip: str) -> None:
    _login_attempts.pop(ip, None)


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    Path(DATABASE_PATH).parent.mkdir(parents=True, exist_ok=True)
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            is_base_user INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS lists (
            id TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            name TEXT NOT NULL,
            type TEXT NOT NULL DEFAULT '',
            date_added TEXT NOT NULL,
            content TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS generations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            trip_type TEXT NOT NULL DEFAULT '',
            location TEXT NOT NULL DEFAULT '',
            duration TEXT NOT NULL DEFAULT '',
            season TEXT NOT NULL DEFAULT '',
            group_size TEXT NOT NULL DEFAULT '',
            weight_priority TEXT NOT NULL DEFAULT '',
            special_considerations TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            markdown TEXT NOT NULL,
            title TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            share_token TEXT UNIQUE
        );
    """)
    if db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        _migrate_from_json(db)
    db.commit()
    db.close()


def _migrate_from_json(db: sqlite3.Connection):
    p = Path(LISTS_FILE)
    if not p.exists():
        return
    try:
        with open(p, "r", encoding="utf-8") as f:
            old_lists = json.load(f)
    except (json.JSONDecodeError, IOError):
        return
    if not old_lists:
        return
    pw_hash = bcrypt.hashpw(BASE_PASS.encode(), bcrypt.gensalt()).decode()
    db.execute(
        "INSERT INTO users (username, password_hash, display_name, is_base_user) VALUES (?, ?, ?, 1)",
        (BASE_USER, pw_hash, BASE_USER.title()),
    )
    uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    for lst in old_lists:
        db.execute(
            "INSERT INTO lists (id, user_id, name, type, date_added, content) VALUES (?, ?, ?, ?, ?, ?)",
            (lst["id"], uid, lst["name"], lst.get("type", ""), lst.get("date_added", date.today().isoformat()), lst["content"]),
        )
    print(f"[migration] Created base user '{BASE_USER}' and imported {len(old_lists)} lists from {LISTS_FILE}")


init_db()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def _check_pw(pw: str, hashed: str) -> bool:
    return bcrypt.checkpw(pw.encode(), hashed.encode())


def _create_token(uid: int, username: str) -> str:
    return jose_jwt.encode(
        {"sub": str(uid), "username": username, "exp": datetime.utcnow() + timedelta(days=30)},
        JWT_SECRET, algorithm="HS256",
    )


def _current_user(request: Request) -> Optional[dict]:
    tok = request.cookies.get("token")
    if not tok:
        return None
    try:
        p = jose_jwt.decode(tok, JWT_SECRET, algorithms=["HS256"])
        return {"id": int(p["sub"]), "username": p["username"]}
    except Exception:
        return None


def _require_user(request: Request) -> dict:
    u = _current_user(request)
    if not u:
        raise HTTPException(status_code=401, detail="Authentication required")
    return u


app.mount("/static", StaticFiles(directory="static"), name="static")


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class RegisterRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50, pattern=r"^[a-zA-Z0-9_-]+$")
    password: str = Field(..., min_length=6, max_length=200)
    display_name: str = Field(..., min_length=1, max_length=100)
    invite_code: str = ""


class LoginRequest(BaseModel):
    username: str
    password: str


class ListCreate(BaseModel):
    name: str = Field(..., max_length=200)
    type: str = Field(..., max_length=100)
    content: str = Field(..., max_length=500_000)


class GenerateRequest(BaseModel):
    trip_type: str = ""
    location: str = ""
    duration: str = ""
    season: str = ""
    group_size: str = ""
    weight_priority: str = ""
    special_considerations: str = ""
    notes: str = ""


class SaveGenerationRequest(BaseModel):
    trip_type: str = ""
    location: str = ""
    duration: str = ""
    season: str = ""
    group_size: str = ""
    weight_priority: str = ""
    special_considerations: str = ""
    notes: str = ""
    markdown: str = ""
    title: str = ""


class ExportRequest(BaseModel):
    markdown: str
    title: str = ""
    group_size: str = ""
    trip_type: str = ""
    location: str = ""
    duration: str = ""
    season: str = ""
    weight_priority: str = ""
    special_considerations: str = ""


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.post("/api/register")
async def register(req: RegisterRequest):
    if INVITE_CODE and not secrets.compare_digest(req.invite_code, INVITE_CODE):
        raise HTTPException(status_code=403, detail="Invalid invite code")
    db = get_db()
    try:
        if db.execute("SELECT 1 FROM users WHERE username = ?", (req.username.lower(),)).fetchone():
            raise HTTPException(status_code=409, detail="Username already taken")
        db.execute(
            "INSERT INTO users (username, password_hash, display_name) VALUES (?, ?, ?)",
            (req.username.lower(), _hash_pw(req.password), req.display_name),
        )
        db.commit()
        uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        resp = JSONResponse({"ok": True, "username": req.username.lower(), "display_name": req.display_name})
        resp.set_cookie("token", _create_token(uid, req.username.lower()), httponly=True, samesite="lax", max_age=30 * 86400)
        return resp
    finally:
        db.close()


@app.post("/api/login")
async def login(req: LoginRequest, request: Request):
    ip = request.client.host if request.client else "unknown"
    _check_rate_limit(ip)
    db = get_db()
    try:
        row = db.execute(
            "SELECT id, username, password_hash, display_name FROM users WHERE username = ?",
            (req.username.lower(),),
        ).fetchone()
        if not row or not _check_pw(req.password, row["password_hash"]):
            _record_failed_login(ip)
            raise HTTPException(status_code=401, detail="Invalid username or password")
        _clear_login_attempts(ip)
        resp = JSONResponse({"ok": True, "username": row["username"], "display_name": row["display_name"]})
        resp.set_cookie("token", _create_token(row["id"], row["username"]), httponly=True, samesite="lax", max_age=30 * 86400)
        return resp
    finally:
        db.close()


@app.post("/api/logout")
async def logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("token")
    return resp


@app.get("/api/me")
async def me(request: Request):
    user = _current_user(request)
    if not user:
        return {"authenticated": False}
    db = get_db()
    try:
        row = db.execute("SELECT display_name, is_base_user FROM users WHERE id = ?", (user["id"],)).fetchone()
        if not row:
            return {"authenticated": False}
        has_lists = db.execute("SELECT COUNT(*) FROM lists WHERE user_id = ?", (user["id"],)).fetchone()[0] > 0
        base_avail = db.execute("SELECT COUNT(*) FROM users WHERE is_base_user = 1 AND id != ?", (user["id"],)).fetchone()[0] > 0
        return {
            "authenticated": True,
            "username": user["username"],
            "display_name": row["display_name"],
            "is_base_user": bool(row["is_base_user"]),
            "has_lists": has_lists,
            "base_library_available": base_avail,
        }
    finally:
        db.close()


@app.post("/api/fork-base-lists")
async def fork_base_lists(request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        base = db.execute("SELECT id FROM users WHERE is_base_user = 1").fetchone()
        if not base:
            return {"copied": 0}
        rows = db.execute("SELECT name, type, date_added, content FROM lists WHERE user_id = ?", (base["id"],)).fetchall()
        for i, r in enumerate(rows):
            db.execute(
                "INSERT INTO lists (id, user_id, name, type, date_added, content) VALUES (?, ?, ?, ?, ?, ?)",
                (str(int(time.time() * 1000) + i), user["id"], r["name"], r["type"], r["date_added"], r["content"]),
            )
        db.commit()
        return {"copied": len(rows)}
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

@app.get("/")
async def root():
    return FileResponse("static/index.html")


@app.get("/share/{token}")
async def share_page(token: str):
    return FileResponse("static/share.html")


# ---------------------------------------------------------------------------
# Lists API
# ---------------------------------------------------------------------------

@app.get("/api/lists")
async def get_lists(request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        return [dict(r) for r in db.execute(
            "SELECT id, name, type, date_added, content FROM lists WHERE user_id = ? ORDER BY date_added DESC",
            (user["id"],),
        ).fetchall()]
    finally:
        db.close()


@app.post("/api/lists")
async def create_list(payload: ListCreate, request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        eid = str(int(time.time() * 1000))
        d = date.today().isoformat()
        db.execute(
            "INSERT INTO lists (id, user_id, name, type, date_added, content) VALUES (?, ?, ?, ?, ?, ?)",
            (eid, user["id"], payload.name.strip(), payload.type.strip(), d, payload.content),
        )
        db.commit()
        return {"id": eid, "name": payload.name.strip(), "type": payload.type.strip(), "date_added": d, "content": payload.content}
    finally:
        db.close()


@app.delete("/api/lists/{list_id}")
async def delete_list(list_id: str, request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        cur = db.execute("DELETE FROM lists WHERE id = ? AND user_id = ?", (list_id, user["id"]))
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="List not found")
        return {"ok": True}
    finally:
        db.close()


@app.post("/api/import-xlsx")
async def import_xlsx(files: List[UploadFile], request: Request):
    user = _require_user(request)
    imported = []
    db = get_db()
    try:
        for upload in files:
            if not upload.filename:
                continue
            raw = await upload.read()
            try:
                wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            except Exception:
                continue
            base_name = Path(upload.filename).stem
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    line = "\t".join(cells).strip()
                    if line and line != "\t" * len(cells):
                        rows_data.append(line)
                if not rows_data:
                    continue
                name = base_name if len(wb.sheetnames) == 1 else f"{base_name} — {sheet_name}"
                content = "\n".join(rows_data)
                lower = base_name.lower()
                trip_type = ""
                for kw, tt in [("climb", "alpine climbing"), ("ski", "ski touring"),
                               ("backpack", "backpacking"), ("hik", "day hiking"),
                               ("raft", "rafting/paddling"), ("river", "rafting/paddling"),
                               ("camp", "car camping"), ("snow camp", "snow camping")]:
                    if kw in lower:
                        trip_type = tt
                        break
                eid = str(int(time.time() * 1000) + len(imported))
                db.execute(
                    "INSERT INTO lists (id, user_id, name, type, date_added, content) VALUES (?, ?, ?, ?, ?, ?)",
                    (eid, user["id"], name[:200], trip_type, date.today().isoformat(), content[:500_000]),
                )
                imported.append({"name": name[:200], "type": trip_type, "items": len(rows_data)})
            wb.close()
        db.commit()
    finally:
        db.close()
    return {"imported": imported, "count": len(imported)}


# ---------------------------------------------------------------------------
# Generate
# ---------------------------------------------------------------------------

@app.post("/api/generate")
async def generate(req: GenerateRequest, request: Request):
    user = _require_user(request)
    if not API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured")

    db = get_db()
    try:
        past_lists = [dict(r) for r in db.execute(
            "SELECT name, type, content FROM lists WHERE user_id = ?", (user["id"],),
        ).fetchall()]
    finally:
        db.close()

    past_lists_block = ""
    if past_lists:
        entries = [f"--- {pl['name']} ({pl['type']}) ---\n{pl['content']}\n" for pl in past_lists]
        past_lists_block = (
            f"\nThe user has provided {len(past_lists)} past packing lists. "
            "Use these to personalize recommendations — reflect what they typically bring, "
            "flag gaps, and note any upgrades worth considering:\n\n" + "\n".join(entries)
        )

    prompt = f"""You are an expert outdoor guide and gear specialist. Generate a comprehensive, \
well-organized packing list for the following trip:

Trip type: {req.trip_type}
Location/objective: {req.location}
Duration: {req.duration}
Season: {req.season}
Group size: {req.group_size}
Weight priority: {req.weight_priority}
Special considerations: {req.special_considerations}
Notes: {req.notes}
{past_lists_block}
Organize by category (Navigation, Shelter, Sleep system, Clothing layers, \
Food & water, Safety, Tools, Personal — use categories appropriate to this \
trip type).

Format each category as a markdown heading (## Category Name) in normal case \
(e.g. "## Navigation & Communication", not "## NAVIGATION & COMMUNICATION") followed by a \
markdown table with these exact columns:

| Item | Priority | Notes |
|------|----------|-------|
| Item name | OPTIONAL | Only if noteworthy |

Rules for the table:
- Priority: leave blank for essential items. Only write OPTIONAL for non-essential items.
- Notes: keep very brief (a few words). Leave blank when the item is self-explanatory. \
No need to explain obvious gear like "pack" or "helmet". Only add a note when there is \
something specific to this trip, season, or group that the user should know.

Be specific to the trip type, season, terrain, and duration.

End with a "Key considerations" section with 2\u20133 specific tips for this \
trip type, location, and season combination. Write Key considerations as \
numbered prose paragraphs, not tables."""

    client = anthropic.Anthropic(api_key=API_KEY)

    def event_stream():
        with client.messages.stream(
            model=MODEL_NAME, max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            for text in stream.text_stream:
                yield f"data: {json.dumps(text)}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Generations API
# ---------------------------------------------------------------------------

@app.post("/api/generations")
async def save_generation(req: SaveGenerationRequest, request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        tok = secrets.token_urlsafe(16)
        db.execute(
            """INSERT INTO generations
               (user_id, trip_type, location, duration, season, group_size,
                weight_priority, special_considerations, notes, markdown, title, share_token)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (user["id"], req.trip_type, req.location, req.duration, req.season,
             req.group_size, req.weight_priority, req.special_considerations,
             req.notes, req.markdown, req.title, tok),
        )
        db.commit()
        gid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": gid, "share_token": tok}
    finally:
        db.close()


@app.get("/api/generations")
async def list_generations(request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        return [dict(r) for r in db.execute(
            """SELECT id, trip_type, location, duration, season, title, created_at, share_token
               FROM generations WHERE user_id = ? ORDER BY created_at DESC""",
            (user["id"],),
        ).fetchall()]
    finally:
        db.close()


@app.get("/api/generations/{gen_id}")
async def get_generation(gen_id: int, request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        row = db.execute(
            """SELECT id, trip_type, location, duration, season, group_size, weight_priority,
                      special_considerations, notes, markdown, title, created_at, share_token
               FROM generations WHERE id = ? AND user_id = ?""",
            (gen_id, user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Not found")
        return dict(row)
    finally:
        db.close()


@app.delete("/api/generations/{gen_id}")
async def delete_generation(gen_id: int, request: Request):
    user = _require_user(request)
    db = get_db()
    try:
        cur = db.execute("DELETE FROM generations WHERE id = ? AND user_id = ?", (gen_id, user["id"]))
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Not found")
        return {"ok": True}
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Share API
# ---------------------------------------------------------------------------

@app.get("/api/share/{token}")
async def get_shared(token: str):
    db = get_db()
    try:
        row = db.execute(
            """SELECT title, markdown, trip_type, location, duration, season,
                      group_size, weight_priority, special_considerations, created_at
               FROM generations WHERE share_token = ?""",
            (token,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Shared list not found")
        return dict(row)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# xlsx export
# ---------------------------------------------------------------------------

def _parse_group_count(s: str) -> int:
    s = s.lower().strip()
    if "solo" in s or s == "1":
        return 1
    m = re.search(r"(\d+)", s)
    if m:
        m2 = re.search(r"(\d+)\s*[-\u2013]\s*(\d+)", s)
        if m2:
            return int(m2.group(2))
        return int(m.group(1))
    return 1


def _parse_markdown_to_rows(text: str) -> list[dict]:
    lines = text.split("\n")
    category = ""
    rows = []
    in_key_considerations = False

    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        header_match = (
            re.match(r"^#{1,3}\s+(.+)", line)
            or re.match(r"^\*\*(.+?)\*\*\s*$", line)
        )
        if header_match:
            h = header_match.group(1).strip()
            if re.search(r"key\s+considerations", h, re.I):
                in_key_considerations = True
                category = ""
                continue
            in_key_considerations = False
            category = h if not h.isupper() else h.title()
            rows.append({"type": "header", "category": category})
            continue
        if in_key_considerations or not category:
            continue
        if re.match(r"^\|[\s\-:]+\|", line):
            continue
        if re.match(r"^\|.*\bItem\b.*\|", line, re.I):
            continue
        if line.startswith("|") and line.endswith("|"):
            cells = [c.strip() for c in line.split("|")[1:-1]]
            if len(cells) < 2:
                continue
            item = re.sub(r"\*\*", "", cells[0]).strip()
            priority = cells[1].strip() if len(cells) > 1 else ""
            note = cells[2].strip() if len(cells) > 2 else ""
            note = re.sub(r"[*_]", "", note)
            if not item or item == "-":
                continue
            rows.append({
                "type": "item",
                "category": category,
                "item": item,
                "priority": priority,
                "note": note,
            })
            continue
        bullet_match = re.match(r"^[-*\d.)\s]+(.+)", line)
        if bullet_match:
            item_text = bullet_match.group(1).strip()
            priority = ""
            if re.search(r"\bOPTIONAL\b", item_text, re.I):
                priority = "OPTIONAL"
                item_text = re.sub(r"\bOPTIONAL\b", "", item_text, flags=re.I)
            item_text = re.sub(r"\*\*", "", item_text).strip()
            parts = re.split(r"\s*[\u2014\u2013\-:]\s+", item_text, maxsplit=1)
            item = parts[0].strip()
            note = parts[1].strip() if len(parts) > 1 else ""
            if item:
                rows.append({
                    "type": "item",
                    "category": category,
                    "item": item,
                    "priority": priority,
                    "note": note,
                })
    return rows


@app.post("/api/export-xlsx")
async def export_xlsx(req: ExportRequest, request: Request):
    _require_user(request)
    rows = _parse_markdown_to_rows(req.markdown)
    person_count = _parse_group_count(req.group_size)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    title_font = Font(name="Calibri", bold=True, size=14)
    detail_font = Font(name="Calibri", size=10, color="444444")
    detail_label_font = Font(name="Calibri", size=10, bold=True, color="444444")
    col_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    item_font = Font(name="Calibri", size=10)
    optional_font = Font(name="Calibri", size=10, color="888888")
    note_font = Font(name="Calibri", size=9, color="666666")
    cat_fill_dark = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cat_fill_light = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
    cat_font = Font(name="Calibri", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    details_parts = []
    if req.duration:
        details_parts.append(req.duration)
    if req.season:
        details_parts.append(f"{req.season} conditions")
    if req.special_considerations:
        details_parts.append(req.special_considerations)
    details_str = ", ".join(details_parts)

    total_cols = 3 + person_count

    r = 1
    title_text = req.title or f"{req.location or 'Packing List'}"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    cell = ws.cell(row=r, column=1, value=title_text)
    cell.font = title_font
    ws.row_dimensions[r].height = 24

    if details_str:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        cell = ws.cell(row=r, column=1, value=f"Trip Details: {details_str}")
        cell.font = detail_font

    meta_fields = [
        ("Route Map:", ""),
        ("Live Tracking:", ""),
        ("Start Location:", ""),
        ("Start Date/Time:", ""),
        ("Estimate Finish Date/Time:", ""),
    ]
    for label, val in meta_fields:
        r += 1
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = detail_label_font

    r += 2

    person_names = ["Person 1"] if person_count == 1 else [f"Person {i+1}" for i in range(person_count)]
    col_headers = ["Category"] + person_names + ["Item", "Essential", "Note"]
    for ci, ch in enumerate(col_headers, 1):
        cell = ws.cell(row=r, column=ci, value=ch)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[r].height = 22
    header_row = r
    r += 1

    current_category = ""
    cat_index = 0

    for row_data in rows:
        if row_data["type"] == "header":
            current_category = row_data["category"]
            cat_index += 1
            continue
        if row_data["type"] != "item":
            continue
        fill = cat_fill_light if cat_index % 2 == 0 else None

        cell = ws.cell(row=r, column=1, value=current_category)
        cell.font = cat_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

        for pi in range(person_count):
            cell = ws.cell(row=r, column=2 + pi)
            cell.alignment = center_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

        item_col = 2 + person_count
        cell = ws.cell(row=r, column=item_col, value=row_data["item"])
        is_optional = row_data["priority"].upper() == "OPTIONAL"
        cell.font = optional_font if is_optional else item_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

        priority_text = row_data["priority"]
        if not priority_text or priority_text.upper() not in ("OPTIONAL",):
            priority_text = "Yes"
        else:
            priority_text = "Optional"
        cell = ws.cell(row=r, column=item_col + 1, value=priority_text)
        cell.font = note_font
        cell.alignment = center_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

        cell = ws.cell(row=r, column=item_col + 2, value=row_data["note"])
        cell.font = note_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

        ws.row_dimensions[r].height = 20
        r += 1

    ws.column_dimensions["A"].width = 28
    for pi in range(person_count):
        ws.column_dimensions[get_column_letter(2 + pi)].width = 10
    item_letter = get_column_letter(2 + person_count)
    ws.column_dimensions[item_letter].width = 38
    ws.column_dimensions[get_column_letter(3 + person_count)].width = 10
    ws.column_dimensions[get_column_letter(4 + person_count)].width = 35

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    last_col_letter = get_column_letter(total_cols + 2)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{r - 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = re.sub(r"[^\w\s\-]", "", req.title or "packing-list").strip()[:80] or "packing-list"
    filename = filename.replace(" ", "_") + ".xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
